# -*- coding: utf-8 -*-
"""
Created on Mon Aug 23 16:32:35 2021

@author: Adrian
"""


import pandas as pd
from openpyxl import load_workbook


pd.options.display.float_format = '{:,.2f}'.format

wb = load_workbook('Bonos en dolares.xlsm',keep_vba=True)


ws = wb["Paridad usd"]

ws.insert_rows(2)





al29 = [['2021-08-25T17:00:11.833', 35.40]]
gd29 = [['2021-08-25T17:00:11.833', 45.40]]
al30 = [['2021-08-25T17:00:11.833', 40.40]]
gd30 = [['2021-08-25T17:00:11.833', 42.40]]
al35 = [['2021-08-25T17:00:11.833', 30.40]]
ae38 = [['2021-08-25T17:00:11.833', 38.40]]
al41 = [['2021-08-25T17:00:11.833', 41.40]]


df1 = pd.DataFrame(al29, columns=('Fecha', 'Precioal29'))
df2 = pd.DataFrame(gd29, columns=('Fecha', 'Preciogd29'))
df3 = pd.DataFrame(al30, columns=('Fecha', 'Precioal30'))
df4 = pd.DataFrame(gd30, columns=('Fecha', 'Preciogd30'))
df5 = pd.DataFrame(al35, columns=('Fecha', 'Precioal35'))
df6 = pd.DataFrame(ae38, columns=('Fecha', 'Precioae38'))
df7 = pd.DataFrame(al41, columns=('Fecha', 'Precioal41'))




filaNueva = ws[2]
filaConFormato = ws[3]

filaConFormato

#formateo las celdas
i = 0
for celda in filaNueva:
    celda._style = filaConFormato[i]._style
    i = i+1 


#df1['Precio'] = df1['Precio'].map('{:,.2f}'.format).str.replace(".", ",",regex=True) 



df1["Fecha"] = pd.to_datetime(
      df1["Fecha"].to_string(index=False)
      ).strftime("%d/%m/%Y")
df2["Fecha"] = pd.to_datetime(
      df2["Fecha"].to_string(index=False)
      ).strftime("%d/%m/%Y")
df3["Fecha"] = pd.to_datetime(
      df3["Fecha"].to_string(index=False)
      ).strftime("%d/%m/%Y")
df4["Fecha"] = pd.to_datetime(
      df4["Fecha"].to_string(index=False)
      ).strftime("%d/%m/%Y")
df5["Fecha"] = pd.to_datetime(
      df5["Fecha"].to_string(index=False)
      ).strftime("%d/%m/%Y")
df6["Fecha"] = pd.to_datetime(
      df6["Fecha"].to_string(index=False)
      ).strftime("%d/%m/%Y")
df7["Fecha"] = pd.to_datetime(
      df7["Fecha"].to_string(index=False)
      ).strftime("%d/%m/%Y")

df1 = pd.merge(df1,df2, on='Fecha')
df1 = pd.merge(df1,df3, on='Fecha')
df1 = pd.merge(df1,df4, on='Fecha')
df1 = pd.merge(df1,df5, on='Fecha')
df1 = pd.merge(df1,df6, on='Fecha')
df1 = pd.merge(df1,df7, on='Fecha')




#    filaNueva[0].value = df1.iloc[0,1]





c= 0

for celda in df1:
    kk = df1.iat[0,c]
    filaNueva[c].value = df1.iloc[0,c]
    c = c+1
    print(celda)    
    

# wb.save('Bonos en dolares.xlsm')







# print (df3)

#df3.to_excel('kkita.xlsx'