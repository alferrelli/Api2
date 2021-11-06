# -*- coding: utf-8 -*-
"""
Created on Sun Aug  8 16:02:08 2021

@author: Adrian
"""

# -*- coding: utf-8 -*-
"""
Created on Thu Jul  8 15:50:13 2021

@author: aferrelli
"""
#import openpyxl
from openpyxl import load_workbook
import requests
import pandas as pd
from openpyxl.formula.translate import Translator
import funcionesIol as fiol
from datetime import datetime


# de estos tickers voy a pedir la cotizacion
simbolos = ['AL29D', 'GD29D', 'AL30D', 'GD30D', 'AL35D', 'AE38D','AL41D']

response_auth = fiol.pedir_token()


if (response_auth.status_code == 200):
        json_data = response_auth.json()

        #aca tengo los dos tokens que necesito
        access_token = json_data['access_token']
        refresh_token = json_data['refresh_token']

        headers = {'Authorization': 'Bearer ' + access_token}

        # pongo los parametros para pedir precio de un tiulo
        # /api/v2/{Mercado}/Titulos/{Simbolo}/Cotizacion
        # simbolo = 'AL29D'
        datosFinales = pd.DataFrame(columns=['fechaHora'])
 
        for simbolo in simbolos:
                
            mercado = 'bCBA'
            plazo = 't2'
   
            parametros = {
                'simbolo': simbolo,
                'mercado': mercado,
                'model.simbolo': simbolo,
                'model.mercado': mercado,
                'model.plazo': plazo
            }

            # Hacemos un Get Request con el Encabezado, que contiene el token de acceso
            # y los parametros de busqueda
            # devuelve el precio actual y fecha (lo que a mi me importa) entre otras cosas
            
            response = requests.get('https://api.invertironline.com/api/v2/{Mercado}/Titulos/{Simbolo}/Cotizacion',
                                     headers=headers,
                                     params=parametros,
                                     verify=True)
            
            if (response.status_code == 200):
                # pedimos los datos en formato JSON
                api_response = response.json()
                df_response = pd.json_normalize(data=api_response)
                df_response = df_response[['fechaHora', 'ultimoPrecio']]
                
                # Como esta abajo guarda la fecha como string
                # df_response['fechaHora'] = pd.to_datetime(df_response["fechaHora"].to_string(index=False)).strftime("%d/%m/%Y")
 
                # Asi guarda la fecha como datetime    
                fecha = pd.to_datetime(df_response["fechaHora"].to_string(index=False)).strftime("%d/%m/%Y")
                df_response['fechaHora'] = datetime.strptime(fecha , "%d/%m/%Y")
                
                datosFinales['fechaHora'] = df_response['fechaHora']
                # agrego la linea de abajo por un warning de columna 'ultimoPrecio' que se repetia
                # pandas le iba poniendo 'ultimoPrecio_x' para diferenciarlos
                datosFinales.rename(columns={"ultimoPrecio":"ultimoPrecio"+simbolo},inplace=True)
                
                datosFinales = pd.merge(datosFinales, df_response, on='fechaHora')
                
                # agrego una columna de paridad asi me quedan las columnas del df
                # igual que las columnas de la planilla de excel
                datosFinales.insert(len(datosFinales.columns) , column="Paridad"+simbolo, value=0)
            else:
                print('Error al obtener contizaciones de: ' + simbolo)
         
            
        wb = load_workbook('Bonos en dolares.xlsm', keep_vba=True)
        ws = wb["Paridad usd"]
        filas = ws.max_row
        rango = "A2:Y"+str(filas)
        ws.move_range(rango,rows=1,translate=True)
        
        filaNueva = ws[2]
        filaConFormato = ws[3]
        
        #formateo las celdas
        i = 0
        for celda in filaNueva:
            celda._style = filaConFormato[i]._style
            i = i+1 
        
        
        # Copio los valores del DF en la nueva celda de Excel
        c = 0
        for column in datosFinales:
            filaNueva[c].value = datosFinales.iloc[0,c]
            c = c+1
        
        # copio las formulas de las columnas con formula a la nueva fila
        col = ['C','E','G','I','K','M','O','P','Q','R','S','T','U','V','W','X','Y']

        for letra_columna in col:
            celda_con_formato = letra_columna+'3'
            celda_nueva = letra_columna+'2'
            
            formula = ws[celda_con_formato].value
            
            
            ws[celda_nueva]=Translator(formula, celda_con_formato).translate_formula(celda_nueva)
        wb.save('Bonos en dolares.xlsm')
        wb.close()
else:
    print ('Error al logearse en la API de IOL')
