# -- coding: utf-8 --
"""
Created on Sun Aug  8 19:54:11 2021

@author: Adrian
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows



wb = load_workbook('Bonos en pesos.xlsx')


#ws = wb["Bonos en pesos"]
ws = wb.active

ws.insert_rows(2)





al29 = [['2021-08-06T17:00:11.833', 35.40]]
al30 = [['2021-08-06T17:00:11.833', 40.40]]


df1 = pd.DataFrame(al29, columns=('fechaHora', 'ultimoPrecio'))
df2 = pd.DataFrame(al30, columns=('fechaHora', 'ultimoPrecio'))



# Formateo fecha


df1['fechaHora'] = pd.to_datetime(
          df1["fechaHora"].to_string(index=False)
          ).strftime("%d/%m/%Y")


df2['fechaHora'] = pd.to_datetime(
          df2["fechaHora"].to_string(index=False)
          ).strftime("%d/%m/%Y")

# join,dos dataframe por fecha



df3 = pd.merge(df1, df2,on='fechaHora')

#df4 = pd.DataFrame(columns=['fechaHora'])

df4=pd.DataFrame()
df4['fechaHora']=df1['fechaHora']
df4 = pd.merge(df4,df3, on='fechaHora')





filaNueva = ws[2]
filaConFormato = ws[3]

filaConFormato

#formateo las celdas
i = 0
for celda in filaNueva:
    celda._style = filaConFormato[i]._style
    i = i+1 


#df1['Precio'] = df1['Precio'].map('{:,.2f}'.format).str.replace(".", ",") 


#filaNueva[0].value = df1["fechaHora"].to_string(index = False)





#convierto a string, quito el index y convierto a float
#filaNueva[1].value = float(df1['ultimoPrecio'].to_string(index=False))

for row in dataframe_to_rows(df3,index=False, header=False): 
    ws.append(row)
    print(row)

wb.save('Bonos en pesos.xlsx')
