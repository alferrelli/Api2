# -*- coding: utf-8 -*-
"""
Created on Sat Oct 16 12:56:12 2021

@author: Adrian
"""


from openpyxl import load_workbook
from datetime import datetime

wb = load_workbook('Bonos en dolares.xlsm', keep_vba=True)
ws = wb["Paridad usd"]

fechas = ws["A"]
i = 0
for celda in fechas:
    if (i>0 and i < 15):
        # fechas[i].value = datetime.strptime(fechas[i].value,"%d/%m/%Y")
        print(i, type(fechas[i].value))
    i = i + 1
    
wb.save('Bonos en dolares.xlsm')
wb.close()

import pandas as pd

wb = pd.read_excel('Bonos en dolares.xlsm',
                   sheet_name="Paridad usd",
                   engine="openpyxl")

kk = wb["Fecha"]