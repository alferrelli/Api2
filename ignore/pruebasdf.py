
"""
Created on Sun Aug  8 19:54:11 2021

@author: Adrian
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

#pd.options.display.float_format = '{:,.2f}'.format

wb = load_workbook('Bonos en dolares.xlsm', keep_vba=True)


ws = wb["Paridad usd"]


titulos = next(ws.values)[0:]

print(titulos)

#ws.insert_rows(2)

filas = ws.max_row
rango = "A2:Y"+str(filas)
ws.move_range(rango,rows=1,translate=True)




al29 = [['2021-08-06T17:00:11.833', 36.15,0]]
gd29 = [['2021-08-06T17:00:11.833', 40.60,0]]
al30 = [['2021-08-06T17:00:11.833', 35.12,0]]
gd30 = [['2021-08-06T17:00:11.833', 38.80,0]]
al35 = [['2021-08-06T17:00:11.833', 33.25,0]]
ae38 = [['2021-08-06T17:00:11.833', 37.75,0]]
al41 = [['2021-08-06T17:00:11.833', 37.40,0]]


df1 = pd.DataFrame(al29, columns=('Fecha', 'Precioal29','Paridad'))
df2 = pd.DataFrame(gd29, columns=('Fecha', 'Preciogd29','Paridad'))
df3 = pd.DataFrame(al30, columns=('Fecha', 'Precioal30','Paridad'))
df4 = pd.DataFrame(gd30, columns=('Fecha', 'Preciogd30','Paridad'))
df5 = pd.DataFrame(al35, columns=('Fecha', 'Precioal35','Paridad'))
df6 = pd.DataFrame(ae38, columns=('Fecha', 'Precioae38','Paridad'))
df7 = pd.DataFrame(al41, columns=('Fecha', 'Precioal41','Paridad'))




filaNueva = ws[2]

filaConFormato = ws[5]



#formateo las celdas
i = 0

for celda in filaNueva:
    celda._style = filaConFormato[i]._style
    i = i+1 


#df1['Precio'] = df1['Precio'].map('{:,.2f}'.format).str.replace(".", ",",regex=True) 



df1["Fecha"] = pd.to_datetime(
      df1["Fecha"].to_string(index=False)
      ).strftime("%d/%m/%Y")
df2["Fecha"] = pd.to_datetime(
      df2["Fecha"].to_string(index=False)
      ).strftime("%d/%m/%Y")
df3["Fecha"] = pd.to_datetime(
      df3["Fecha"].to_string(index=False)
      ).strftime("%d/%m/%Y")
df4["Fecha"] = pd.to_datetime(
      df4["Fecha"].to_string(index=False)
      ).strftime("%d/%m/%Y")
df5["Fecha"] = pd.to_datetime(
      df5["Fecha"].to_string(index=False)
      ).strftime("%d/%m/%Y")
df6["Fecha"] = pd.to_datetime(
      df6["Fecha"].to_string(index=False)
      ).strftime("%d/%m/%Y")
df7["Fecha"] = pd.to_datetime(
      df7["Fecha"].to_string(index=False)
      ).strftime("%d/%m/%Y")

df1 = pd.merge(df1,df2, on='Fecha')
df1 = pd.merge(df1,df3, on='Fecha')
df1 = pd.merge(df1,df4, on='Fecha')
df1 = pd.merge(df1,df5, on='Fecha')
df1 = pd.merge(df1,df6, on='Fecha')
df1 = pd.merge(df1,df7, on='Fecha')

# Copio los valores del DF en la nueva celda de Excel
c = 0
for column in df1:
    filaNueva[c].value = df1.iloc[0,c]
    c = c+1


col = ['C','E','G','I','K','M','O','P','Q','R','S','T','U','V','W','X','Y']

for letra_columna in col:
    celda_con_formato = letra_columna+'3'
    celda_nueva = letra_columna+'2'
    
    formula = ws[celda_con_formato].value
    ws[celda_nueva]=Translator(formula, celda_con_formato).translate_formula(celda_nueva)
    

formula = ws['C3'].value
nuevaF = Translator(formula, "C3").translate_formula("C2")
ws['C2'] = nuevaF


wb.save('Bonos en dolares.xlsm')







# print (df3)

#df3.to_excel('kkita.xlsx')